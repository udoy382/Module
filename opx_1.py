# Tech With Tim / Automate Excel With Python- Python Excel Tutorial (OpenPyXL)

'''
# import openpyxl
from openpyxl import Workbook, load_workbook, workbook

# wb = workbook
# wb = load_workbook('grades.xlsx')

# ws = worksheet
# ws = wb.active

# ws['A2'].value = 'Test'  # chnage value on the sheet

# print(ws) # show hwo amy sheet on this xl file
# print(ws['A2'].value) # show value on this A2 colom or row

# wb.save('grades.xlsx') # if im change any value on the sheet so must used this function, and before save file closed the xl file

# print(wb.sheetnames) # show all sheet on the file

# ws = wb['Sheet1']
# print(wb['Sheet1'])
'''

#       creat new sheet and update or save sheet on the xl file
'''
wb.create_sheet('Test')
print(wb.sheetnames)
wb.save('grades.xlsx')
'''

#-------------------------------------------

# from openpyxl import Workbook, load_workbook
# from openpyxl.utils import get_column_letter

#       append values on xl sheet 
'''
wb = Workbook()
ws = wb.active
ws.title = 'Data'

ws.append(['Udoy', 'Is', 'Great', '!'])
ws.append(['Maryam', 'Is', 'Great', '!'])
ws.append(['Fariha', 'Is', 'Great', '!'])
ws.append(['Amy', 'Is', 'Great', '!'])
wb.save('grades.xlsx')
'''

#       read 1 to 10 row and 1 to 4 colum
'''
wb = load_workbook('grades.xlsx')
ws = wb.active

for row in range(1, 11):
    for col in range(1, 5):
        char =get_column_letter(col)
        print(ws[char + str(row)].value)

wb.save('grades.xlsx')
'''

#       merge and unmerge, giving row or colum A1 to D1 or more
'''
wb = load_workbook('grades.xlsx')
ws = wb.active

# ws.merge_cells('A1:D1')
ws.merge_cells('A1:D2') # merge two row
# ws.unmerge_cells('A1:D1')
wb.save('grades.xlsx')
'''
#       make blank rows or colum, insert any new new line
'''
wb = load_workbook('grades.xlsx')
ws = wb.active

# ws.insert_rows(7)
# ws.delete_rows(7)

# ws.insert_cols(2)
ws.delete_cols(2)

wb.save('grades.xlsx')
'''

#       move item excel sheet to another sheet
'''
wb = load_workbook('grades.xlsx')
ws = wb.active

ws.move_range('C1:D11', rows=2, cols=2)

wb.save('grades.xlsx')
'''

#       add data on new xl sheet
'''
from openpyxl import Workbook, load_workbook, workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Udoy": {
        'math':46,
        'english':72,
        'science':88,
        'ict':100,
        'gym':77
    },
    "Maryam": {
        'math':56,
        'english':34,
        'science':65,
        'ict':98,
        'gym':88
    },
    "Fariha": {
        'math':65,
        'english':56,
        'science':56,
        'ict':45,
        'gym':90
    },
    "Amy": {
        'math':98,
        'english':67,
        'science':67,
        'ict':45,
        'gym':55
    }
}

wb = Workbook()
ws = wb.active
ws.title = 'Grades'


headings = ['name'] + list(data['Udoy'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

#       sum of the value
for col in range(2, len(data['Udoy']) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

#       chnage color of row or col
for col in range(1, 7):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color='0099CCFF')

wb.save('grades.xlsx')
'''
