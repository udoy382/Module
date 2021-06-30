# from web

from openpyxl import *
import datetime

'''
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3, 4])

# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save('info.xlsx')
'''
#----------------------------------------
'''
workbook = Workbook()
sheet = workbook.active

sheet['A1'] = 'Hello'
sheet['B1'] = 'World'

workbook.save(filename='info.xlsx')
'''
#-------------------------------------------