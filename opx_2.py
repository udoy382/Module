# ForMyScholars / #33 Writing/Updating Operations on Excel file Using Openpyxl Module

import openpyxl as op

#       Creating an excel file
'''
z = op.Workbook() # make a new excel file
z.save('C:/Users/User/OneDrive/Documents/Module/b.xlsx') # set locatinon and name where save excel file
'''

#       work something with excel file
'''
z = op.Workbook()
print(z.sheetnames) # find excel sheets name
s = z['Sheet'] # select excel sheet for change name
s.title = 'Udoy' # change sheets name on excel file

z.create_sheet() # creat new sheets on excel file
z.create_sheet() # creat one new sheets
z.create_sheet(title='Hola') # creat new sheets with name

#       copy excel file, copy everything b.xlsx to a.xlsh , we first work b.xlsx the copy all a.xlsx
# z.save('C:/Users/User/OneDrive/Documents/Module/a.xlsx')
# z.save('C:/Users/User/OneDrive/Documents/Module/b.xlsx')
'''

#       writing or updating the values

z = op.Workbook()
s = z['Sheet']

# s.cell(row=1, column=2, value='Saifur Rahman Udoy') # writing the excel file
s.cell(row=5, column=2, value='20') # add some value excel file
s.cell(row=5, column=2, value='40') # update value, 20 to 40

s.merge_cells('A6:B6') # merge meen blank line seperet positon,single rows
s.unmerge_cells('A6:B6') # unmerge the cells

s.cell(row=6, column=1).value='Hey Hacker' # write value on giving position 

#       append grupe of value on excel file with append function, append funcion add value botum of the list,
f = ((1, 2, 3), (3, 4, 5), (5, 6, 7))
for g in f:
    s.append(g) 

s['C11'].value="=SUM(A7:C9)" # sum all appending value

z.save('C:/Users/User/OneDrive/Documents/Module/b.xlsx') # save every changes